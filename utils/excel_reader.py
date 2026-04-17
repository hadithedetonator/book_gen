"""
utils/excel_reader.py
----------------------
Reads book input rows from a local Excel (.xlsx) file using openpyxl.

The workbook must contain a sheet named exactly "Books" (see EXCEL_SHEET_NAME).
The first row is treated as the header row; column order does not matter —
columns are resolved by name using the header row values.

Required columns:
  title                   — book title (non-empty required)
  notes_on_outline_before — editorial notes before outline generation

Optional columns (may be empty):
  notes_on_outline_after  — reviewer notes after initial outline
  status_outline_notes    — one of: yes | no | no_notes_needed
"""

import logging
from pathlib import Path
from typing import Iterator

import openpyxl

from book_gen.constants import EXCEL_SHEET_NAME, ExcelColumn

log = logging.getLogger(__name__)


class ExcelReadError(Exception):
    """Raised when the Excel file cannot be read or is malformed."""


def _resolve_headers(sheet) -> dict[str, int]:
    """
    Build a column-name → 0-based index map from the first row of *sheet*.

    Args:
        sheet: An openpyxl worksheet object.

    Returns:
        Dict mapping lowercase column names to their 0-based column indices.

    Raises:
        ExcelReadError: If the header row is empty.
    """
    header_row = [
        (cell.value or "").strip().lower() for cell in next(sheet.iter_rows(max_row=1))
    ]
    if not any(header_row):
        raise ExcelReadError("The 'Books' sheet has an empty header row.")
    return {name: idx for idx, name in enumerate(header_row) if name}


def _get_cell(row: tuple, col_map: dict[str, int], column: str) -> str:
    """
    Safely retrieve and clean a cell value from a row tuple.

    Args:
        row:     Tuple of cell objects for one row.
        col_map: Column-name → index map from _resolve_headers.
        column:  The column name (case-insensitive, will be lowercased).

    Returns:
        Stripped string value, or "" if the column is absent or the cell is None.
    """
    idx = col_map.get(column.lower())
    if idx is None or idx >= len(row):
        return ""
    value = row[idx].value
    return str(value).strip() if value is not None else ""


def read_books(file_path: str) -> Iterator[dict]:
    """
    Yield one dict per data row in the 'Books' sheet of the given Excel file.

    Empty title rows are skipped with a warning. All rows — including those
    with missing optional fields — are yielded so that gate logic in Stage 1
    can emit the appropriate error event.

    Args:
        file_path: Absolute or relative path to the .xlsx file.

    Yields:
        Dict with keys:
          title (str), notes_on_outline_before (str),
          notes_on_outline_after (str), status_outline_notes (str)

    Raises:
        ExcelReadError: If the file does not exist, is not a valid workbook,
                        or the 'Books' sheet is missing.
    """
    path = Path(file_path)
    if not path.exists():
        raise ExcelReadError(f"Excel file not found: {file_path}")
    if path.suffix.lower() != ".xlsx":
        raise ExcelReadError(f"Expected a .xlsx file, got: {path.suffix}")

    log.info("Opening Excel file: %s", file_path)

    try:
        workbook = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReadError(f"Cannot open workbook '{file_path}': {exc}") from exc

    if EXCEL_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise ExcelReadError(
            f"Sheet '{EXCEL_SHEET_NAME}' not found in '{file_path}'. "
            f"Available sheets: {workbook.sheetnames}"
        )

    sheet = workbook[EXCEL_SHEET_NAME]
    rows = list(sheet.iter_rows())          # consume generator before close
    workbook.close()

    if len(rows) < 2:
        log.warning("'Books' sheet has no data rows (only a header or is empty).")
        return

    col_map = _resolve_headers(sheet.__class__)   # header already consumed; rebuild from rows[0]
    # Re-derive col_map from rows[0] directly
    col_map = {
        (cell.value or "").strip().lower(): idx
        for idx, cell in enumerate(rows[0])
        if (cell.value or "").strip()
    }

    for row_idx, row in enumerate(rows[1:], start=2):  # 1-indexed for user messages
        title = _get_cell(row, col_map, ExcelColumn.TITLE)
        if not title:
            log.warning("Row %d skipped: 'title' column is empty.", row_idx)
            continue

        record = {
            ExcelColumn.TITLE:                   title,
            ExcelColumn.NOTES_ON_OUTLINE_BEFORE:  _get_cell(row, col_map, ExcelColumn.NOTES_ON_OUTLINE_BEFORE),
            ExcelColumn.NOTES_ON_OUTLINE_AFTER:   _get_cell(row, col_map, ExcelColumn.NOTES_ON_OUTLINE_AFTER),
            ExcelColumn.STATUS_OUTLINE_NOTES:     _get_cell(row, col_map, ExcelColumn.STATUS_OUTLINE_NOTES),
        }
        log.info("Read row %d: title='%s'", row_idx, title)
        yield record
